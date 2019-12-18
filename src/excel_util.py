import pathlib

import pandas as pd
from colour import Color
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series, StockChart
from openpyxl.chart.axis import ChartLines, DateAxis
from openpyxl.chart.data_source import NumData, NumVal
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.chart.updown_bars import UpDownBars
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink, HyperlinkList

from src.log import print_exception
from src.columns import CHART_COLUMNS, UPPER_RANGE, LOWER_RANGE

def create_worksheet(ew, df, name, file_name, index=1):
    df.to_excel(excel_writer=ew, sheet_name=name)
    ws = ew.book[name]
    cels = [f"A{x}" for x in range(2, len(df) + 3)]
    for cl in cels:
        ws[cl].style = "custom_datetime1"
    cels = [f"B{x}" for x in range(2, len(df) + 3)]
    for cl in cels:
        ws[cl].style = "custom_datetime1"
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 11
    # Helps to navigate to parent sheet
    hyp = f"{file_name}#'SUMMARY'!B{index + 2}"
    ws["A1"].hyperlink = hyp
    ws["A1"].hyperlink.location = hyp.split("#")[1]
    ws["A1"].style = "custom_datetime"


def create_summary_sheet(ew, df, file_name):
    df.to_excel(excel_writer=ew, sheet_name="SUMMARY")
    ws = ew.book["SUMMARY"]
    disps = [f"{file_name}#'{x:%Y-%m-%d}'!A1" for x in df.ED]
    cels = [f"B{x}" for x in range(2, len(disps) + 3)]
    for cl, hyp in zip(cels, disps):
        ws[cl].hyperlink = hyp
        ws[cl].hyperlink.location = hyp.split("#")[1]
        ws[cl].style = "custom_datetime"
    cels = [f"A{x}" for x in range(2, len(disps) + 3)]
    for cl in cels:
        ws[cl].style = "custom_datetime1"
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 11

def create_inputsheet(ew, df):
    df.to_excel(excel_writer=ew, sheet_name="INPUTS")
    ws = ew.book["INPUTS"]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

def add_style(ew):
    ns = NamedStyle(name="custom_datetime", number_format="YYYY-MM-DD")
    ns.font = Font(underline="single", color="0000FF")
    ew.book.add_named_style(ns)
    ns = NamedStyle(name="custom_datetime1", number_format="YYYY-MM-DD")
    ns.font = Font(color="0000FF")
    ew.book.add_named_style(ns)

def create_line_series(ws, min_col, min_row, max_row, labels, color, legend_loc=0):
    l2 = LineChart()
    l2.add_data(
        Reference(ws, min_col=min_col, min_row=min_row, max_row=max_row),
        titles_from_data=True,
    )
    l2.set_categories(labels)
    l2.series[0].graphicalProperties.line.solidFill = color
    s1 = l2.series[0]
    s1.dLbls = DataLabelList()
    # Initialize data label
    dl = DataLabel()
    # Set properties
    dl.showVal = True
    dl.showSerName = True
    dl.idx = legend_loc
    # position t for top
    dl.position = "r"
    # Append data label to data lebels
    s1.dLbls.dLbl.append(dl)
    return l2
#
def create_work_sheet_chart(ew, df, title, name):
    #
    df.to_excel(excel_writer=ew, sheet_name=name)
    ws = ew.book[name]
    #
    dfl = len(df) + 1
    #
    labels = Reference(ws, min_col=1, min_row=2, max_row=dfl)
    #
    ost = df.columns.get_loc("OPEN") + 2
    cnd = df.columns.get_loc("CLOSE") + 2
    #
    l1 = StockChart()
    data = Reference(ws, min_col=ost, max_col=cnd, min_row=1, max_row=dfl)
    l1.add_data(data, titles_from_data=True)
    l1.set_categories(labels)
    #
    for s in l1.series:
        s.graphicalProperties.line.noFill = True
    #
    l1.hiLowLines = ChartLines()
    l1.upDownBars = UpDownBars()
    if title is not None:
        l1.title = title
    # add dummy cache
    pts = [NumVal(idx=i) for i in range(len(data) - 1)]
    cache = NumData(pt=pts)
    l1.series[-1].val.numRef.numCache = cache
    #
    if dfl <= 6:
        l1.height = 15
        l1.width = 7
    elif dfl >= 6 and dfl <= 25:
        l1.height = 15
        # l1.width = 30
        l1.width = 10
    else:
        l1.height = 20
        # l1.width = 40
        l1.width = 10
    # Monthly constant sigma lines
    clen = len(CHART_COLUMNS)
    #
    colors = list(Color("#ff4554").range_to(Color("#ffc7cb"), clen))
    colors = [x.get_hex()[1:] for x in colors]
    #
    def create_lines(cols, l1, loc):
        try:
            sli = df.columns.get_loc(cols[0]) + 2
            sln = sli + len(cols)
            for i, xy in enumerate(range(sli, sln)):
                l1 += create_line_series(
                    ws, xy, 1, dfl, labels, colors[i], legend_loc=loc
                )
        except Exception as e:
            print_exception(e)
            print(f"Unable to plot given cols")

    # Creats monthly sigma cols
    create_lines(CHART_COLUMNS, l1, 0)
    #
    mn = df[LOWER_RANGE].min() - 100
    mx = df[UPPER_RANGE].max() + 100
    l1.x_axis.number_format = "yyyymmmdd"
    l1.y_axis.scaling.min = mn
    l1.y_axis.scaling.max = mx
    l1.y_axis.majorUnit = 200
    l1.legend = None
    ws.add_chart(l1, "A2")
    ws.column_dimensions["A"].width = 11
    for cell in ws["A"]:
        cell.style = "custom_datetime"
    ws.column_dimensions["K"].width = 11
    for cell in ws["K"]:
        cell.style = "custom_datetime"
    return ws